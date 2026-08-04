from __future__ import annotations

import json
from dataclasses import replace
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableCell, TableRow
from odf.text import P
from openpyxl import Workbook

from nan_fung.datasources import common
from nan_fung.datasources.common import AcquisitionResponse, StoredAcquisitionResponse
from nan_fung.ingestion.file_release_lifecycle import (
    FileReleaseLifecycleError,
    acquire_live_file_release,
    ingest_file_release_artifacts,
    reparse_file_release_evidence,
)
from nan_fung.ingestion.registry import DatasourceRegistry, default_registry
from nan_fung.ingestion.file_release_workflow import (
    EPC_DATASOURCE_ID,
    HYBRID_DATASOURCE_ID,
    VOA_DATASOURCE_ID,
    FileReleaseWorkflowError,
    contract_for,
)
from nan_fung.operational import OperationalError, OperationalStore
from nan_fung.storage.db import connect_database


def _voa_zip() -> bytes:
    output = BytesIO()
    with ZipFile(output, "w") as archive:
        header = "geography,area_code,area_name,2025,2026\n"
        archive.writestr(
            "table_SOP5_1.csv",
            header + "REGL,E12000007,London,100,103400\n",
        )
        archive.writestr(
            "table_SOP5_2.csv",
            header + "REGL,E12000007,London,900,9264908\n",
        )
    return output.getvalue()


def _hybrid_xlsx() -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Table_6"
    worksheet.cell(row=1, column=1, value="Trends in hybrid working")
    worksheet.cell(row=2, column=1, value="Great Britain, 20 March 2020 to 28 June 2026")
    worksheet.cell(
        row=10,
        column=1,
        value=(
            "Percentage of working adults that have both travelled to work and "
            "worked from home in the past seven days"
        ),
    )
    worksheet.cell(row=10, column=2, value="All\npersons\n%")
    worksheet.cell(row=10, column=3, value="All\npersons\nLCL")
    worksheet.cell(row=10, column=4, value="All\npersons\nUCL")
    worksheet.cell(row=11, column=1, value="3 to 28 June 2026")
    worksheet.cell(row=11, column=2, value=25)
    worksheet.cell(row=11, column=3, value=22)
    worksheet.cell(row=11, column=4, value=28)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _add_row(table: Table, values: list[object]) -> None:
    row = TableRow()
    for value in values:
        if isinstance(value, (int, float)):
            cell = TableCell(valuetype="float", value=value)
        else:
            cell = TableCell(valuetype="string")
            cell.addElement(P(text=str(value)))
        row.addElement(cell)
    table.addElement(row)


def _epc_ods() -> bytes:
    document = OpenDocumentSpreadsheet()
    table = Table(name="A_by_Region")
    _add_row(
        table,
        [
            "A- Non-Domestic Properties by Region by Energy Performance Asset "
            "Rating - in each Year/Quarter to 30 June 2026"
        ],
    )
    _add_row(table, ["This worksheet contains one table."])
    _add_row(
        table,
        [
            "Source: Energy Performance Certificates for Buildings Register for "
            "England and Wales"
        ],
    )
    _add_row(
        table,
        [
            "Region",
            "Quarter",
            "Number Lodgements",
            "Total Floor Area (m2)",
            "A+",
            "A",
            "B",
            "C",
            "D",
            "E",
            "F",
            "G",
            "Not Recorded",
        ],
    )
    _add_row(
        table,
        ["London", "2026/2", 3630, 3102511, 13, 482, 1621, 1019, 358, 119, 13, 5, 0],
    )
    document.spreadsheet.addElement(table)
    output = BytesIO()
    document.save(output)
    return output.getvalue()


def _response(url: str, body: bytes, media_type: str) -> AcquisitionResponse:
    return AcquisitionResponse(
        request_url=url,
        final_url=url,
        status=200,
        headers={"Content-Type": media_type},
        body=body,
        retrieved_at="2026-08-01T00:00:00Z",
        method="GET",
    )


def _artifacts(datasource_id: str) -> tuple[AcquisitionResponse, AcquisitionResponse]:
    contract = contract_for(datasource_id)
    if datasource_id == VOA_DATASOURCE_ID:
        release_url = (
            "https://assets.publishing.service.gov.uk/media/example/"
            "ndr_stock_of_properties_2026.zip"
        )
        discovery = _response(
            contract.discovery_url,
            f'<a href="{release_url}">release</a>'.encode(),
            "text/html",
        )
        return discovery, _response(release_url, _voa_zip(), "application/zip")
    if datasource_id == HYBRID_DATASOURCE_ID:
        release_url = (
            "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/wellbeing/"
            "datasets/publicopinionsandsocialtrendsgreatbritainworkingarrangements/"
            "3to28june2026/workingarrangements3to28june2026.xlsx"
        )
        discovery = _response(
            contract.discovery_url,
            f'<a href="{release_url}">release</a>'.encode(),
            "text/html",
        )
        return discovery, _response(
            release_url,
            _hybrid_xlsx(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    assert datasource_id == EPC_DATASOURCE_ID
    release_url = "https://assets.publishing.service.gov.uk/media/example/table-a.ods"
    discovery = _response(
        contract.discovery_url,
        json.dumps(
            {
                "details": {
                    "attachments": [
                        {
                            "title": (
                                "Table A: Non-domestic Energy Performance Certificates by energy "
                                "performance asset rating"
                            ),
                            "url": release_url,
                        }
                    ]
                }
            }
        ).encode(),
        "application/json",
    )
    return discovery, _response(
        release_url,
        _epc_ods(),
        "application/vnd.oasis.opendocument.spreadsheet",
    )


@pytest.mark.parametrize(
    ("datasource_id", "expected_source_date"),
    (
        (VOA_DATASOURCE_ID, "2026-03-31"),
        (HYBRID_DATASOURCE_ID, "2026-06-03"),
        (EPC_DATASOURCE_ID, "2026-06-30"),
    ),
)
def test_file_release_fixture_lifecycle_persists_two_artifacts_and_promotes(
    tmp_path: Path, datasource_id: str, expected_source_date: str
) -> None:
    store = OperationalStore(tmp_path)
    discovery, release = _artifacts(datasource_id)

    result = ingest_file_release_artifacts(
        store,
        datasource_id,
        discovery=discovery,
        release=release,
    )

    assert result.status == "succeeded"
    assert result.canonical_changed is True
    assert result.discovery_evidence_id is not None
    assert result.observation_ids
    connection = connect_database(store.database_path, read_only=True)
    try:
        roles = connection.execute(
            "SELECT role FROM run_evidence WHERE run_id = ? ORDER BY role",
            (result.run_id,),
        ).fetchall()
        locator = connection.execute(
            "SELECT locator_json FROM observation_evidence WHERE observation_id = ?",
            (result.observation_ids[0],),
        ).fetchone()
        canonical = connection.execute(
            "SELECT observation_id, source_date FROM canonical_latest_v1 WHERE observation_id = ?",
            (result.observation_ids[0],),
        ).fetchone()
    finally:
        connection.close()
    assert [row["role"] for row in roles] == ["discovery", "primary"]
    assert locator is not None
    record_locator = json.loads(locator["locator_json"])["record_locator"]
    if datasource_id == EPC_DATASOURCE_ID:
        assert record_locator["row"] == 5
    assert canonical is not None
    assert canonical["source_date"] == expected_source_date


def test_file_release_reparse_inherits_nonproduction_lane(tmp_path: Path) -> None:
    store = OperationalStore(tmp_path)
    discovery, release = _artifacts(VOA_DATASOURCE_ID)
    original = ingest_file_release_artifacts(
        store,
        VOA_DATASOURCE_ID,
        discovery=discovery,
        release=release,
        lane="source_discovery",
    )

    replay = reparse_file_release_evidence(
        store, VOA_DATASOURCE_ID, original.evidence_id
    )

    assert replay.canonical_changed is False
    connection = connect_database(store.database_path, read_only=True)
    try:
        lineage = connection.execute(
            """
            SELECT j.trigger, j.job_kind, j.request_json
            FROM ingestion_run AS r
            JOIN workflow_job AS j ON j.job_id = r.job_id
            WHERE r.run_id = ?
            """,
            (replay.run_id,),
        ).fetchone()
    finally:
        connection.close()
    assert lineage is not None
    assert lineage["trigger"] == "reparse"
    assert lineage["job_kind"] == "offline_reparse"
    assert json.loads(lineage["request_json"]) == {
        "reparse_evidence_id": original.evidence_id
    }
    with pytest.raises(OperationalError, match="must match the source lane"):
        reparse_file_release_evidence(
            store,
            VOA_DATASOURCE_ID,
            original.evidence_id,
            lane="production_ingestion",
        )


def test_file_release_stored_release_does_not_write_its_cas_object_twice(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = OperationalStore(tmp_path)
    discovery, release = _artifacts(HYBRID_DATASOURCE_ID)
    stored_discovery = store.artifacts.put_bytes(
        discovery.body, media_type="text/html"
    )
    stored_release = store.artifacts.put_bytes(
        release.body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    monkeypatch.setattr(
        store.artifacts,
        "put_bytes",
        lambda *_args, **_kwargs: pytest.fail("stored file releases must not be rewritten"),
    )

    result = ingest_file_release_artifacts(
        store,
        HYBRID_DATASOURCE_ID,
        discovery=StoredAcquisitionResponse(
            request_url=discovery.request_url,
            final_url=discovery.final_url,
            status=discovery.status,
            headers=discovery.headers,
            artifact=stored_discovery,
            retrieved_at=discovery.retrieved_at,
            method=discovery.method,
        ),
        release=StoredAcquisitionResponse(
            request_url=release.request_url,
            final_url=release.final_url,
            status=release.status,
            headers=release.headers,
            artifact=stored_release,
            retrieved_at=release.retrieved_at,
            method=release.method,
        ),
    )

    assert result.status == "succeeded"


def test_live_file_release_capture_streams_discovery_and_release_to_cas(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    store = OperationalStore(tmp_path)
    contract = contract_for(VOA_DATASOURCE_ID)
    selection_url = (
        "https://www.gov.uk/government/statistics/"
        "non-domestic-rating-stock-of-properties-march-2026"
    )
    release_url = (
        "https://assets.publishing.service.gov.uk/media/example/"
        "ndr_stock_of_properties_2026.zip"
    )
    opened: list[str] = []
    gate_events: list[tuple[str, bool]] = []

    class Gate:
        def permit(self, host: str, *, continuation: bool = False) -> None:
            gate_events.append((host, continuation))

        def record_response(
            self, _host: str, *, status: int, retry_after: str | None
        ) -> None:
            assert status == 200
            assert retry_after is None
            return None

    class Response:
        status = 200

        def __init__(self, url: str, body: bytes, media_type: str) -> None:
            self.headers = {"Content-Type": media_type}
            self._url = url
            self._chunks = iter((body[:7], body[7:]))

        def __enter__(self) -> "Response":
            return self

        def __exit__(self, *_args: object) -> None:
            return None

        def close(self) -> None:
            return None

        def read(self, _size: int) -> bytes:
            return next(self._chunks, b"")

        def getcode(self) -> int:
            return self.status

        def geturl(self) -> str:
            return self._url

    def open_once(request, _timeout: int) -> Response:
        opened.append(request.full_url)
        if request.full_url == contract.discovery_url:
            return Response(
                request.full_url,
                f'<a href="{selection_url}">release</a>'.encode(),
                "text/html",
            )
        if request.full_url == selection_url:
            return Response(
                request.full_url,
                f'<a href="{release_url}">release</a>'.encode(),
                "text/html",
            )
        assert request.full_url == release_url
        return Response(request.full_url, _voa_zip(), "application/zip")

    monkeypatch.setattr(common, "_open_once", open_once)
    monkeypatch.setattr(
        store.artifacts,
        "put_bytes",
        lambda *_args, **_kwargs: pytest.fail("live release capture must stream to CAS"),
    )

    capture = acquire_live_file_release(
        VOA_DATASOURCE_ID,
        artifact_store=store.artifacts,
        host_gate=Gate(),
        resolver=lambda _host: ("8.8.8.8",),
    )

    assert opened == [contract.discovery_url, selection_url, release_url]
    assert gate_events == [
        ("www.gov.uk", False),
        ("www.gov.uk", True),
        ("assets.publishing.service.gov.uk", False),
    ]
    assert isinstance(capture.discovery, StoredAcquisitionResponse)
    assert isinstance(capture.selection, StoredAcquisitionResponse)
    assert isinstance(capture.release, StoredAcquisitionResponse)
    assert store.artifacts.verify(capture.discovery.artifact)
    assert capture.selection is not None
    assert store.artifacts.verify(capture.selection.artifact)
    assert store.artifacts.verify(capture.release.artifact)


def test_voa_release_selection_persists_a_three_artifact_lineage(tmp_path: Path) -> None:
    store = OperationalStore(tmp_path)
    contract = contract_for(VOA_DATASOURCE_ID)
    selection_url = (
        "https://www.gov.uk/government/statistics/"
        "non-domestic-rating-stock-of-properties-march-2026"
    )
    release_url = (
        "https://assets.publishing.service.gov.uk/media/example/"
        "ndr_stock_of_properties_2026.zip"
    )
    discovery = _response(
        contract.discovery_url,
        f'<a href="{selection_url}">March 2026</a>'.encode(),
        "text/html",
    )
    selection = _response(
        selection_url,
        f'<a href="{release_url}">download</a>'.encode(),
        "text/html",
    )
    release = _response(release_url, _voa_zip(), "application/zip")

    result = ingest_file_release_artifacts(
        store,
        VOA_DATASOURCE_ID,
        discovery=discovery,
        selection=selection,
        release=release,
    )

    assert result.status == "succeeded"
    assert result.discovery_evidence_id is not None
    assert result.selection_evidence_id is not None
    connection = connect_database(store.database_path, read_only=True)
    try:
        roles = connection.execute(
            "SELECT role FROM run_evidence WHERE run_id = ? ORDER BY role",
            (result.run_id,),
        ).fetchall()
    finally:
        connection.close()
    assert [row["role"] for row in roles] == ["discovery", "primary", "supporting"]


def test_file_release_refuses_a_release_url_not_selected_by_its_saved_discovery(
    tmp_path: Path,
) -> None:
    store = OperationalStore(tmp_path)
    discovery, release = _artifacts(VOA_DATASOURCE_ID)
    forged = AcquisitionResponse(
        request_url="https://assets.publishing.service.gov.uk/media/other/ndr_stock_of_properties_2026.zip",
        final_url="https://assets.publishing.service.gov.uk/media/other/ndr_stock_of_properties_2026.zip",
        status=200,
        headers=release.headers,
        body=release.body,
        retrieved_at=release.retrieved_at,
        method="GET",
    )

    with pytest.raises(FileReleaseWorkflowError, match="request URL does not match"):
        ingest_file_release_artifacts(
            store, VOA_DATASOURCE_ID, discovery=discovery, release=forged
        )


def test_file_release_refuses_a_definition_with_changed_executable_bindings(
    tmp_path: Path,
) -> None:
    seed = default_registry()
    version_one = seed.lookup(VOA_DATASOURCE_ID)
    changed = replace(version_one, definition_version=2, parser_version="v2")
    store = OperationalStore(
        tmp_path,
        registry=DatasourceRegistry((version_one, changed), seed.sources),
    )
    discovery, release = _artifacts(VOA_DATASOURCE_ID)

    with pytest.raises(FileReleaseLifecycleError, match="new bound lifecycle"):
        ingest_file_release_artifacts(
            store,
            VOA_DATASOURCE_ID,
            discovery=discovery,
            release=release,
            definition_version=2,
        )

    assert not (tmp_path / "evidence").exists()
