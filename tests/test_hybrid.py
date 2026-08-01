from io import BytesIO

import pytest
from openpyxl import Workbook

from nan_fung.datasources import hybrid


def _hybrid_workbook(*, valid_schema: bool = True) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Table_6"
    if valid_schema:
        worksheet.cell(row=1, column=1, value="Trends in hybrid working")
        worksheet.cell(row=2, column=1, value="Great Britain, 20 March 2020 to 28 June 2026")
        worksheet.cell(
            row=10,
            column=1,
            value=(
                "Percentage of working adults that have both travelled to work and "
                "worked from home in the past seven days"
            ),
        )
        worksheet.cell(row=10, column=2, value="All\npersons\n%")
        worksheet.cell(row=10, column=3, value="All\npersons\nLCL")
        worksheet.cell(row=10, column=4, value="All\npersons\nUCL")
    worksheet.cell(row=11, column=1, value="6 to 31 May 2026")
    worksheet.cell(row=11, column=2, value=28)
    worksheet.cell(row=11, column=3, value=26)
    worksheet.cell(row=11, column=4, value=31)
    worksheet.cell(row=12, column=1, value="3 to 28 June 2026")
    worksheet.cell(row=12, column=2, value=25)
    worksheet.cell(row=12, column=3, value=22)
    worksheet.cell(row=12, column=4, value=28)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_fetch_hybrid_working_returns_latest_proxy(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(hybrid, "get_bytes", lambda _url, **_kwargs: _hybrid_workbook())

    result = hybrid.fetch_hybrid_working()

    assert result["category"] == "hybrid_working"
    assert result["published_at"] == "2026-07-17"
    assert [row["period"] for row in result["records"]] == [
        "6 to 31 May 2026",
        "3 to 28 June 2026",
    ]
    assert result["records"][-1]["estimate_percent"] == 25
    assert result["records"][-1]["indicator_type"] == "proxy"
    assert result["records"][-1]["is_office_occupancy"] is False


def test_parse_hybrid_working_xlsx_is_a_pure_artifact_parser() -> None:
    records = hybrid.parse_hybrid_working_xlsx(_hybrid_workbook())

    assert [record["period"] for record in records] == [
        "6 to 31 May 2026",
        "3 to 28 June 2026",
    ]


def test_parse_hybrid_working_xlsx_rejects_an_unrecognized_table_schema() -> None:
    with pytest.raises(ValueError, match="schema is not recognized"):
        hybrid.parse_hybrid_working_xlsx(_hybrid_workbook(valid_schema=False))


def test_parse_hybrid_working_dataset_html_selects_one_official_workbook() -> None:
    assert hybrid.parse_hybrid_working_dataset_html(
        b'<a href="/file?uri=/datasets/workingarrangements/july/'
        b'workingarrangementsjuly2027.xlsx">download</a>'
        b'<a href="/file?uri=/datasets/other/other.xlsx">other</a>'
    ) == (
        "https://www.ons.gov.uk/file?uri=/datasets/workingarrangements/july/"
        "workingarrangementsjuly2027.xlsx"
    )


def test_parse_hybrid_working_dataset_html_refuses_unsafe_or_nonmatching_links() -> None:
    for href in (
        "https://example.test/file?uri=/datasets/workingarrangements/july/working.xlsx",
        "https://www.ons.gov.uk/file?uri=/datasets/other/other.xlsx",
    ):
        with pytest.raises(ValueError):
            hybrid.parse_hybrid_working_dataset_html(
                f'<a href="{href}">download</a>'.encode()
            )


def test_parse_hybrid_working_dataset_html_refuses_ambiguous_workbooks() -> None:
    with pytest.raises(ValueError, match="ambiguous"):
        hybrid.parse_hybrid_working_dataset_html(
            b'<a href="/file?uri=/datasets/workingarrangements/june/'
            b'workingarrangementsjune2027.xlsx">old</a>'
            b'<a href="/file?uri=/datasets/workingarrangements/july/'
            b'workingarrangementsjuly2027.xlsx">new</a>'
        )


def test_discover_hybrid_working_xlsx_url(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        hybrid,
        "get_bytes",
        lambda _url, **_kwargs: (
            b'<a href="/file?uri=/workingarrangements/july/workingarrangements.xlsx">'
            b"download</a>"
        ),
    )

    assert hybrid.discover_hybrid_working_xlsx_url("https://ons.example/dataset") == (
        "https://ons.example/file?uri=/workingarrangements/july/workingarrangements.xlsx"
    )


@pytest.mark.network
@pytest.mark.legacy_live_probe
def test_fetch_hybrid_working_live() -> None:
    result = hybrid.fetch_hybrid_working()

    assert len(result["records"]) == 2
    record = result["records"][-1]
    assert record["period"] == "3 to 28 June 2026"
    assert record["estimate_percent"] == 25
    assert record["lower_confidence_limit"] <= record["estimate_percent"]
    assert record["estimate_percent"] <= record["upper_confidence_limit"]
    assert record["is_office_occupancy"] is False
