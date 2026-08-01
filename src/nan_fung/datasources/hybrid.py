"""Free ONS hybrid-working indicators for Great Britain."""

from __future__ import annotations

import html
import re
from html.parser import HTMLParser
from io import BytesIO
from urllib.parse import parse_qsl, urljoin, urlparse, urlunparse

from openpyxl import load_workbook

from nan_fung.datasources.common import SourceResult, get_bytes, source_result
from nan_fung.ingestion.policies import ArtifactPolicy, SourcePolicy, validate_zip_artifact

ONS_WORKING_ARRANGEMENTS_URL = (
    "https://www.ons.gov.uk/peoplepopulationandcommunity/wellbeing/datasets/"
    "publicopinionsandsocialtrendsgreatbritainworkingarrangements"
)
ONS_JUNE_2026_XLSX_URL = (
    "https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/wellbeing/"
    "datasets/publicopinionsandsocialtrendsgreatbritainworkingarrangements/"
    "3to28june2026/workingarrangements3to28june2026.xlsx"
)
_HYBRID_ARTIFACT_POLICY = ArtifactPolicy(max_bytes=250 * 1024 * 1024)
_ONS_POLICY = SourcePolicy(("www.ons.gov.uk",), allowed_query_keys=("uri",))
_ONS_BASE_URL = "https://www.ons.gov.uk"
_HYBRID_TITLE_PREFIX = "trends in hybrid working"
_HYBRID_GEOGRAPHY_PREFIX = "great britain,"
_HYBRID_HEADERS = (
    "percentage of working adults that have both travelled to work and worked from home in the past seven days",
    "all persons %",
    "all persons lcl",
    "all persons ucl",
)


def fetch_hybrid_working() -> SourceResult:
    """Return the last two numeric ONS hybrid-working estimates from June 2026.

    This is a Great Britain survey proxy, not London office occupancy or
    building-access data.
    """

    observations = parse_hybrid_working_xlsx(
        get_bytes(ONS_JUNE_2026_XLSX_URL, policy=_ONS_POLICY)
    )
    return source_result(
        category="hybrid_working",
        source="ONS Opinions and Lifestyle Survey (OPN)",
        source_url=ONS_JUNE_2026_XLSX_URL,
        published_at="2026-07-17",
        records=observations[-2:],
    )


def parse_hybrid_working_xlsx(evidence: bytes) -> list[dict[str, object]]:
    """Parse a captured ONS workbook without acquiring a new edition."""

    validate_zip_artifact(evidence, _HYBRID_ARTIFACT_POLICY)
    workbook = load_workbook(
        BytesIO(evidence), read_only=True, data_only=True, keep_links=False
    )
    try:
        worksheet = workbook["Table_6"]
        _validate_hybrid_working_schema(worksheet)
        observations = []
        for row_number, (period, estimate, lower, upper) in enumerate(
            worksheet.iter_rows(min_row=11, max_col=4, values_only=True),
            start=11,
        ):
            if not _is_number(estimate):
                continue
            if (
                not isinstance(period, str)
                or not period.strip()
                or not _is_number(lower)
                or not _is_number(upper)
                or not 0 <= lower <= estimate <= upper <= 100
            ):
                raise ValueError("ONS hybrid workbook has an invalid Table_6 row")
            observations.append(
                {
                    "period": period,
                    "geography": "Great Britain",
                    "metric": (
                        "working adults who both travelled to work and worked "
                        "from home in the past seven days"
                    ),
                    "estimate_percent": estimate,
                    "lower_confidence_limit": lower,
                    "upper_confidence_limit": upper,
                    "source_row": row_number,
                    "indicator_type": "proxy",
                    "is_office_occupancy": False,
                }
            )
        return observations
    except KeyError as error:
        raise ValueError("ONS hybrid workbook has no Table_6 sheet") from error
    finally:
        workbook.close()


def _validate_hybrid_working_schema(worksheet: object) -> None:
    title = _normalise_cell(worksheet.cell(1, 1).value)  # type: ignore[attr-defined]
    geography = _normalise_cell(worksheet.cell(2, 1).value)  # type: ignore[attr-defined]
    headers = tuple(
        _normalise_cell(worksheet.cell(10, column).value)  # type: ignore[attr-defined]
        for column in range(1, 5)
    )
    if (
        not title.startswith(_HYBRID_TITLE_PREFIX)
        or not geography.startswith(_HYBRID_GEOGRAPHY_PREFIX)
        or headers != _HYBRID_HEADERS
    ):
        raise ValueError("ONS hybrid workbook Table_6 schema is not recognized")


def _normalise_cell(value: object) -> str:
    return " ".join(value.split()).casefold() if isinstance(value, str) else ""


def _is_number(value: object) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def discover_hybrid_working_xlsx_url(
    dataset_url: str = ONS_WORKING_ARRANGEMENTS_URL,
) -> str:
    """Find a current ONS workbook release from the stable dataset page."""

    evidence = get_bytes(dataset_url, policy=_ONS_POLICY)
    if dataset_url == ONS_WORKING_ARRANGEMENTS_URL:
        return parse_hybrid_working_dataset_html(evidence)

    # Retain the historical injected-dataset adapter for existing callers.
    page = evidence.decode("utf-8", errors="replace")
    candidates = re.findall(r'''href=["']([^"']+\.xlsx(?:\?[^"']*)?)["']''', page, re.I)
    matches = [
        urljoin(dataset_url, html.unescape(candidate))
        for candidate in candidates
        if "workingarrangements" in candidate.lower()
        or "working-arrangements" in candidate.lower()
    ]
    if not matches:
        raise ValueError("no hybrid-working XLSX was found on the dataset page")
    return matches[-1]


class _HrefCollector(HTMLParser):
    """Collect anchor href values without treating script text as page links."""

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.hrefs: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.casefold() != "a":
            return
        for name, value in attrs:
            if name.casefold() == "href" and value:
                self.hrefs.append(value)


def parse_hybrid_working_dataset_html(evidence: bytes) -> str:
    """Select one safe working-arrangements XLSX from captured ONS HTML."""

    candidates: set[str] = set()
    for href in _hrefs_from_html(evidence):
        url = urljoin(_ONS_BASE_URL, html.unescape(href).strip())
        if not _looks_like_hybrid_working_xlsx_url(url):
            continue
        if not _is_approved_hybrid_working_xlsx_url(url):
            raise ValueError("ONS working-arrangements XLSX URL is not approved")
        candidates.add(_canonical_hybrid_working_xlsx_url(url))
    if not candidates:
        raise ValueError("no approved ONS working-arrangements XLSX was found in dataset HTML")
    if len(candidates) != 1:
        raise ValueError("ONS dataset HTML contains ambiguous working-arrangements XLSX links")
    return candidates.pop()


def _hrefs_from_html(evidence: bytes) -> tuple[str, ...]:
    parser = _HrefCollector()
    parser.feed(evidence.decode("utf-8", errors="replace"))
    parser.close()
    return tuple(parser.hrefs)


def _looks_like_hybrid_working_xlsx_url(url: str) -> bool:
    parsed = urlparse(url)
    if parsed.path != "/file":
        return False
    query = parse_qsl(parsed.query, keep_blank_values=True)
    if len(query) != 1 or query[0][0] != "uri":
        return False
    return _is_working_arrangements_xlsx_uri(query[0][1])


def _is_approved_hybrid_working_xlsx_url(url: str) -> bool:
    try:
        parsed = urlparse(url)
        return (
            parsed.scheme.casefold() == "https"
            and parsed.hostname == "www.ons.gov.uk"
            and parsed.username is None
            and parsed.password is None
            and parsed.port is None
            and not parsed.fragment
            and _looks_like_hybrid_working_xlsx_url(url)
        )
    except ValueError:
        return False


def _is_working_arrangements_xlsx_uri(uri: str) -> bool:
    parsed = urlparse(uri)
    if (
        parsed.scheme
        or parsed.netloc
        or parsed.query
        or parsed.fragment
        or not parsed.path.startswith("/")
        or "\\" in parsed.path
        or "/../" in f"/{parsed.path.strip('/')}"
    ):
        return False
    normalized_path = re.sub(r"[-_]", "", parsed.path.casefold())
    return normalized_path.endswith(".xlsx") and "workingarrangements" in normalized_path


def _canonical_hybrid_working_xlsx_url(url: str) -> str:
    parsed = urlparse(url)
    return urlunparse(("https", "www.ons.gov.uk", "/file", "", parsed.query, ""))
