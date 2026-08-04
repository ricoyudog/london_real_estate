"""Test-only canonical hybrid-working seeder; not shipped runtime logic."""

from __future__ import annotations

import argparse
from datetime import UTC, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook

from nan_fung.ingestion.file_release_workflow import record_metadata_for
from nan_fung.operational import OperationalStore


RETRIEVED_AT = datetime(2026, 8, 1, 12, tzinfo=UTC)
DATASOURCE_ID = "ons.opn.hybrid_working"
SOURCE_URL = "https://www.ons.gov.uk/file?uri=/workingarrangements.xlsx"
RECORD = {"period": "3 to 28 June 2026", "geography": "Great Britain", "metric": "working adults who both travelled to work and worked from home in the past seven days", "estimate_percent": "25", "lower_confidence_limit": "22", "upper_confidence_limit": "28", "source_row": 11, "indicator_type": "proxy", "is_office_occupancy": False}


def _evidence_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Table_6"
    sheet.cell(1, 1, "Trends in hybrid working")
    sheet.cell(2, 1, "Great Britain, 3 to 28 June 2026")
    sheet.cell(10, 1, "Percentage of working adults that have both travelled to work and worked from home in the past seven days")
    sheet.cell(10, 2, "All persons %")
    sheet.cell(10, 3, "All persons LCL")
    sheet.cell(10, 4, "All persons UCL")
    sheet.append(["3 to 28 June 2026", 25, 22, 28])
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def seed_hybrid_working(data_dir: Path) -> None:
    """Seed a Great Britain ONS hybrid-working proxy observation."""

    store = OperationalStore(data_dir)
    store.sync_registry(now=RETRIEVED_AT)
    metadata = record_metadata_for(DATASOURCE_ID, RECORD)
    queued = store.enqueue(DATASOURCE_ID, scheduled_for=RETRIEVED_AT, request={"fixture": "agent-runtime-hybrid-working"}, trigger="manual")
    claim = store.claim_job(queued.job_id, "fixture-worker", now=RETRIEVED_AT)
    if claim is None:
        raise RuntimeError("fixture job could not be claimed")
    run = store.start_run(claim, "fixture-worker", now=RETRIEVED_AT)
    evidence = store.persist_evidence(run, _evidence_bytes(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", request={"method": "GET", "url": SOURCE_URL}, response={"status": 200, "final_url": SOURCE_URL}, source_id="ons.opn", retrieved_at=RETRIEVED_AT, now=RETRIEVED_AT)
    store.persist_observation(run, record_key=metadata["record_key"], payload=RECORD, record_type=metadata["record_type"], category="hybrid_working", evidence=(evidence,), source_date=metadata["source_date"], period_label=metadata["period_label"], unit=metadata["unit"], definition_text=metadata["definition"], limitations=metadata["limitations"], locator=metadata["locator"], now=RETRIEVED_AT)
    store.finish_run(run, status="succeeded", promote=True, now=RETRIEVED_AT)


def main() -> None:
    parser = argparse.ArgumentParser(description="Seed a canonical hybrid-working fixture")
    parser.add_argument("data_dir", type=Path)
    args = parser.parse_args()
    seed_hybrid_working(args.data_dir)


if __name__ == "__main__":
    main()
